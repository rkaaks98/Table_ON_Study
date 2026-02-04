# src/clients/easypos_kiosk_reader.py
import psycopg2
import requests
import time
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# --- 설정 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# config와 recipe 경로는 상위 디렉토리를 기준으로 재설정합니다.
CONFIG_PATH = os.path.join(BASE_DIR, '..', '..', '..', 'config', 'config.json')
RECIPE_PATH = os.path.join(BASE_DIR, '..', '..', '..', 'config', 'recipe.json')
LOG_DIR = os.path.join(BASE_DIR, '..', '..', '..', 'logs') # 로그 저장 디렉토리
LAST_DATETIME_FILE = os.path.join(BASE_DIR, 'easypos_last_datetime.txt') # 키오스크별 파일 분리
ORDER_SERVER_URL = 'http://localhost:8100' # 주문 서버 URL 고정

# --- 함수 정의 ---

def log_orders_to_excel(orders_list):
    """주문 리스트를 엑셀 파일에 일괄 기록합니다."""
    if not orders_list:
        return

    try:
        if not os.path.exists(LOG_DIR):
            os.makedirs(LOG_DIR)

        today_str = datetime.now().strftime("%Y-%m-%d")
        log_file = os.path.join(LOG_DIR, f"{today_str}_orders.xlsx")
        
        headers = ["접수일시", "주문번호", "메뉴코드", "메뉴명", "수량"]
        
        if not os.path.exists(log_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Order Log"
            ws.append(headers)
            for i, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = 15
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['D'].width = 25
        else:
            wb = openpyxl.load_workbook(log_file)
            ws = wb.active

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for order in orders_list:
            # order: (order_no, menu_code, menu_name, quantity)
            new_row = [timestamp] + list(order)
            ws.append(new_row)
        
        wb.save(log_file)
        print(f"[LOG] {len(orders_list)} orders logged to {log_file}")

    except PermissionError:
        print(f"[LOG][ERROR] 엑셀 파일이 열려있어 로그를 저장할 수 없습니다: {log_file}")
    except Exception as e:
        print(f"[LOG][ERROR] Failed to log order reception to Excel: {e}")

def load_db_config():
    """설정 파일에서 easyPOS DB 정보를 로드합니다."""
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            config = json.load(f)
        kiosk_config = config.get('kiosk', {})
        if kiosk_config.get('brand') == 'easypos':
            return kiosk_config.get('easypos')
        return None
    except Exception as e:
        print(f"[ERROR] 설정 파일 로드/분석 실패: {e}")
        return None

def get_menu_info():
    """레시피 파일에서 메뉴 코드와 메뉴 이름을 매핑한 딕셔너리를 반환합니다."""
    try:
        with open(RECIPE_PATH, 'r', encoding='utf-8') as f:
            recipes = json.load(f)
        return {r['menu_code']: r['menu_name'] for r in recipes}
    except Exception as e:
        print(f"[ERROR] 레시피 파일 처리 실패: {e}")
        return {}

def get_last_processed_time():
    """마지막 처리 시간을 파일에서 읽어옵니다."""
    try:
        with open(LAST_DATETIME_FILE, 'r') as f:
            return f.read().strip()
    except FileNotFoundError:
        return '19700101000000000001'

def save_last_processed_time(datetime_str):
    """마지막 처리 시간을 파일에 저장합니다."""
    try:
        with open(LAST_DATETIME_FILE, 'w') as f:
            f.write(datetime_str)
    except IOError as e:
        print(f"[ERROR] 마지막 주문 시간 저장 실패: {e}")

def fetch_and_process_orders(db_config, menu_info):
    """DB에서 신규 주문을 가져와 서버로 전송하고 엑셀에 기록하는 메인 로직."""
    last_time = get_last_processed_time()
    valid_menu_codes = menu_info.keys()
    
    try:
        with psycopg2.connect(**db_config) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM FUNC_SALE_DATE_VW('01') as (sale_date character varying, bill_qty numeric, employ_code character varying);")
                sale_date_result = cur.fetchone()
                if not sale_date_result or not sale_date_result[0]:
                    print("[INFO] 현재 영업일이 아닙니다.")
                    return
                
                opendate = sale_date_result[0]

                query = f"SELECT * FROM vw_ord_label_print_header WHERE sale_date = '{opendate}' AND create_datetime > '{last_time}' ORDER BY create_datetime;"
                cur.execute(query)
                new_bills = cur.fetchall()

                latest_time_in_batch = last_time
                orders_to_log = []

                for bill in new_bills:
                    bill_no, create_datetime = bill[2], bill[4]
                    
                    detail_query = f"SELECT * FROM vw_ord_label_print_detail WHERE sale_date = '{opendate}' AND bill_no = '{bill_no}';"
                    cur.execute(detail_query)
                    
                    for detail in cur.fetchall():
                        order_no, menu_code, quantity = int(detail[3]), int(detail[5]), int(detail[7])
                        
                        if menu_code in valid_menu_codes:
                            menu_name = menu_info.get(menu_code, "Unknown Menu")
                            print(f"[Order] 신규 주문 발견: 주문번호={order_no}, 메뉴={menu_name}, 수량={quantity}")
                            
                            # 서버 전송 시도
                            try:
                                for _ in range(quantity):
                                    requests.get(f"{ORDER_SERVER_URL}/addOrder/{order_no}/{menu_code}", timeout=5).raise_for_status()
                                    time.sleep(0.2)
                                
                                # 전송 성공 시에만 로그 리스트에 추가
                                orders_to_log.append((order_no, menu_code, menu_name, quantity))

                            except requests.RequestException as req_err:
                                print(f"[API ERROR] 주문 전송 실패 (주문번호={order_no}): {req_err}")
                                # 실패 시 로그에 남기지 않고, DB 타임스탬프도 갱신하지 않아야 함 (다음 루프에서 재시도)

                        else:
                            print(f"[SKIP] 판매하지 않는 메뉴: 주문번호={order_no}, 메뉴코드={menu_code}")
                    
                    latest_time_in_batch = create_datetime
                
                # 일괄 로그 저장
                if orders_to_log:
                    log_orders_to_excel(orders_to_log)

                if latest_time_in_batch != last_time:
                    save_last_processed_time(latest_time_in_batch)

    except psycopg2.Error as db_err:
        print(f"[DB ERROR] DB 작업 중 오류: {db_err}")
    except requests.RequestException as api_err:
        print(f"[API ERROR] 주문 서버 통신 오류: {api_err}")


if __name__ == "__main__":
    print("--- EasyPOS 키오스크 주문 리더 시작 ---")
    db_config = load_db_config()
    if not db_config:
        print("[FATAL] EasyPOS DB 설정을 찾을 수 없어 프로그램을 종료합니다.")
        exit(1)

    while True:
        try:
            menu_info = get_menu_info()
            if menu_info:
                fetch_and_process_orders(db_config, menu_info)
            else:
                print("[WARN] 유효한 메뉴 정보를 찾을 수 없습니다. 10초 후 재시도합니다.")
                time.sleep(10)

        except Exception as e:
            print(f"[UNEXPECTED ERROR] 예상치 못한 오류 발생: {e}")
            time.sleep(10)
        
        print("...다음 주문 대기중...")
        time.sleep(1)

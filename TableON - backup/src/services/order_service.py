import time
import threading
import requests
import json
import os
import copy
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.utils import get_column_letter

from enum import Enum, auto
from typing import List, Dict, Optional, Any
from queue import Queue, Empty

import logging
from logging.handlers import TimedRotatingFileHandler

# --- Logger Setup ---
LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs')
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)

# Order Service Logger
logger = logging.getLogger("OrderService")
logger.setLevel(logging.INFO)

log_handler = TimedRotatingFileHandler(
    os.path.join(LOG_DIR, "order_service_daily.log"),
    when="midnight",
    interval=1,
    backupCount=30,
    encoding='utf-8'
)
log_formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
log_handler.setFormatter(log_formatter)
logger.addHandler(log_handler)

console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
logger.addHandler(console_handler)

ROBOT_SERVICE_URL = "http://localhost:8300"
IO_SERVICE_URL = "http://localhost:8400"
DEVICE_SERVICE_URL = "http://localhost:8500"
PICKUP_SERVICE_URL = "http://localhost:8600"

CONFIG_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'config')
RECIPE_PATH = os.path.join(CONFIG_DIR, 'recipe.json')

SIMULATION_MODE = False
COFFEE_BRAND = ""
PICKUP_MODE = "sensor"  # "sensor" or "rotate"

# --- 써모플랜 보일러 온도 보상 로직 설정 ---
IDLE_TIME_THRESHOLD_SECONDS = 5 * 60  # 5분
EXTRA_DURATION_SECONDS = 20

try:
    with open(os.path.join(CONFIG_DIR, 'config.json'), 'r', encoding='utf-8') as f:
        config_data = json.load(f)
        SIMULATION_MODE = config_data.get('simulation_mode', False)
        COFFEE_BRAND = config_data.get('coffee_machine', {}).get('brand', '')
        PICKUP_MODE = config_data.get('pickup_mode', 'rotate')
        print(f"[Config] Simulation Mode: {SIMULATION_MODE}")
        print(f"[Config] Coffee Brand: {COFFEE_BRAND}")
        print(f"[Config] Pickup Mode: {PICKUP_MODE}")
except Exception as e:
    print(f"[Config] Failed to load config.json: {e}")


REG_CMD       = 600   # FSM_CMD: 모션 커맨드 (PC → 로봇)
REG_INIT      = 700   # FSM_INIT: 모션 커맨드 완료 (로봇 → PC)
REG_FSM_STAT  = 900   # FSM_STAT: 인디 동작상태 (0: 대기, 1: 동작)

REG_CUP_IDX    = 100   # cup_idx: 컵 index (1: 핫, 2: 아이스)
REG_PICKUP_IDX = 101   # pickup_idx: 픽업대 슬롯 index
REG_CUP_RES    = 102   # cup_result: 컵 추출 성공 여부 (1: 성공, 2: 실패)
REG_CUP_SET    = 103   # cup_set:  컵D 컵 거치 완료 (1: 완료)
REG_CUP_MOVE   = 104   # cup_move : 컵D 센서로 이동 (1: 이동 완료)
REG_CUP_SENSOR = 105   # CUP_SENSOR: 컵 센서 (1: ON, 2: OFF)
REG_CUP_ON     = 106   # CUP_ON: 컵 도착신호 (1: 도착)
REG_SYRUP_IDX  = 107   # SYRUP_IDX: 시럽 종류 (1~8)

REG_F_PICK_IDX = REG_PICKUP_IDX # 픽업대대

CMD_CUP_MOVE     = 110  # HOME → 컵디스펜서 N point → 610

CMD_WI_MOVE      = 111  # 제빙기 N point → 제빙기 추출구 → 611
CMD_WI_DONE      = 112  # 제빙기 추출구 → 제빙기 N point → 612

CMD_COFFEE_MOVE  = 113  # 커피머신 N point → 커피머신 추출구 → 613
CMD_COFFEE_DONE  = 114  # 커피머신 추출구 → 커피머신 N point → 614
CMD_COFFEE_PLACE = 115  # 커피머신에 컵 거치 (병렬 처리용) → 615
CMD_COFFEE_PICK  = 116  # 커피머신에서 컵 픽업 (병렬 처리용) → 616

CMD_HOT_MOVE     = 117  # 온수기 N point → 온수기 추출구 → 617
CMD_HOT_DONE     = 118  # 온수기 추출구 → 온수기 N point → 618

CMD_PICKUP_MOVE  = 119  # 픽업대 N 포인트 이동 → 619
CMD_PICKUP_PLACE = 120  # 서빙 (LED 점등) pickup_idx → 620

CMD_SYRUP_MOVE   = 121  # 시럽 N 포인트 → 시럽 추출구 → 621 (SYRUP_IDX)
CMD_SYRUP_DONE   = 122  # 시럽 추출구 → 시럽 N 포인트 → 622

CMD_HOME         = 123  # 픽업대 N 포인트 → 홈 → 623

CMD_BREATHING    = 200  # 홈 → 홈 → 700

CMD_DESC = {
    CMD_CUP_MOVE:    "HOME -> CUP (컵 배출)",
    CMD_WI_MOVE:     "제빙기 접근 (얼음/물/탄산수)",
    CMD_WI_DONE:     "제빙기 완료",
    CMD_COFFEE_MOVE: "커피머신 접근",
    CMD_COFFEE_DONE: "커피머신 완료",
    CMD_COFFEE_PLACE:"커피머신 컵 거치 (병렬)",
    CMD_COFFEE_PICK: "커피머신 컵 픽업 (병렬)",
    CMD_HOT_MOVE:    "온수기 접근",
    CMD_HOT_DONE:    "온수기 완료",
    CMD_PICKUP_MOVE: "픽업대 접근",
    CMD_PICKUP_PLACE:"픽업대 서빙",
    CMD_SYRUP_MOVE:  "시럽 접근",
    CMD_SYRUP_DONE:  "시럽 완료",
    CMD_HOME:        "홈 복귀",
    CMD_BREATHING:   "BREATHING (대기)"
}

# System Modes (무인 모드만 사용)
MODE_MANUAL   = 0
MODE_AUTO     = 1

ORDER_WAITING    = "WAITING"
ORDER_PROCESSING = "PROCESSING"
ORDER_COMPLETED  = "COMPLETED"
ORDER_CANCELLED  = "CANCELLED"

NODERED_URL = "http://localhost:1880/notify"

def notify_clients(event_name, data=None):
    payload = {'event': event_name}
    if data:
        payload.update(data)
    
    def _send():
        try:
            requests.post(NODERED_URL, json=payload, timeout=0.5)
        except:
            pass
            
    threading.Thread(target=_send, daemon=True).start()

def log_performance_to_excel(order_info):
    try:
        if not os.path.exists(LOG_DIR):
            os.makedirs(LOG_DIR)

        today_str = datetime.now().strftime("%Y-%m-%d")
        log_file = os.path.join(LOG_DIR, f"{today_str}_system_performance.xlsx")
        
        headers = ["접수일시", "완료일시", "주문번호", "메뉴코드", "메뉴명", "상태"]
        
        if not os.path.exists(log_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "System Performance"
            ws.append(headers)
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['E'].width = 25
        else:
            wb = openpyxl.load_workbook(log_file)
            ws = wb.active

        created_dt = datetime.fromtimestamp(order_info['created_at']).strftime("%Y-%m-%d %H:%M:%S")
        completed_dt = datetime.fromtimestamp(order_info['completed_at']).strftime("%Y-%m-%d %H:%M:%S")
        
        new_row = [
            created_dt,
            completed_dt,
            order_info['order_no'],
            order_info['menu_code'],
            order_info['menu_name'],
            order_info['status']
        ]
        ws.append(new_row)
        wb.save(log_file)

    except PermissionError:
        logger.error(f"[Performance] 엑셀 파일이 열려있어 저장이 불가능합니다: {log_file}")
    except Exception as e:
        logger.error(f"[Performance] Failed to log to Excel: {e}")


class RobotInterface:
    """Robot Service 통신 래퍼 (HTTP)"""
    def __init__(self, robot_id: str):
        self.robot_id = robot_id
        self.base_url = ROBOT_SERVICE_URL

    def get_status(self) -> Optional[Dict]:
        try:
            res = requests.get(f"{self.base_url}/status/{self.robot_id}", timeout=10)
            if res.status_code == 200:
                return res.json()
        except:
            pass
        return None

    def write_register(self, addr: int, value: int) -> bool:
        try:
            res = requests.post(f"{self.base_url}/writeRegister", 
                              json={"robot_id": self.robot_id, "addr": addr, "value": value}, timeout=10)
            return res.status_code == 200
        except Exception as e:
            print(f"[{self.robot_id}] Write Reg Fail: {e}")
            return False

    def read_register(self, addr: int) -> int:
        try:
            res = requests.post(f"{self.base_url}/readRegister", 
                              json={"robot_id": self.robot_id, "addr": addr}, timeout=20)
            if res.status_code == 200:
                val = res.json().get('value')
                if val is not None:
                    return int(val)
        except Exception as e:
            print(f"[{self.robot_id}] Read Reg {addr} Error: {e}")
        return -1

    def send_command(self, cmd_code: int) -> bool:
        return self.write_register(REG_CMD, cmd_code)
    
    def wait_init(self, target_val: int, timeout=600.0) -> bool:
        """Wait until REG_INIT becomes target_val"""
        t0 = time.time()
        print(f"[{self.robot_id}] Waiting Init: Target={target_val} (Timeout={timeout}s)")
        
        while time.time() - t0 < timeout:
            if system_mode != MODE_AUTO:
                print(f"[{self.robot_id}] Wait Init Aborted: System switched to Manual Mode")
                return False
                
            val = self.read_register(REG_INIT)
            if val == target_val:
                print(f"[{self.robot_id}] Init Matched: {val}")
                return True
                
            time.sleep(0.5)
            
        print(f"[{self.robot_id}] Wait Init Timeout!")
        return False


class DeviceInterface:
    def __init__(self):
        self.base_url = DEVICE_SERVICE_URL

    def make_coffee(self, product_id, duration):
        """커피 추출 명령 전송 (응답 대기 없음)"""
        def _send():
            try:
                requests.get(f"{self.base_url}/coffee/{product_id}/{duration}", timeout=180)
            except Exception as e:
                print(f"[Device] Coffee Error: {e}")
        threading.Thread(target=_send, daemon=True).start()
        print(f"[Device] Coffee command sent: product_id={product_id}, duration={duration}")
        return True

    def make_coffee_async(self, product_id, duration):
        """비동기 커피 추출 시작 (병렬 처리용) - make_coffee와 동일"""
        self.make_coffee(product_id, duration)

    def execute_rinse(self):
        """린스 명령 전송 (응답 대기 없음)"""
        def _send():
            try:
                requests.get(f"{self.base_url}/coffee/rinse", timeout=60)
            except Exception as e:
                print(f"[Device] Rinse Error: {e}")
        threading.Thread(target=_send, daemon=True).start()
        print("[Device] Rinse command sent")
        return True

    def dispense_ice_water(self, ice_time, water_time):
        try:
            res = requests.get(f"{self.base_url}/waterice/{ice_time}/{water_time}", timeout=180)
            return res.status_code == 200
        except Exception as e:
            print(f"[Device] Ice/Water Error: {e}")
            return False

    def dispense_syrup(self, code, duration):
        try:
            res = requests.get(f"{self.base_url}/syrup/{code}/{duration}", timeout=180)
            return res.status_code == 200
        except Exception as e:
            print(f"[Device] Syrup Error: {e}")
            return False

    def dispense_hot_water(self, duration):
        try:
            res = requests.get(f"{self.base_url}/hotwater/{duration}", timeout=180)
            return res.status_code == 200
        except Exception as e:
            print(f"[Device] HotWater Error: {e}")
            return False
            
    def dispense_sparkling(self, duration):
        try:
            res = requests.get(f"{self.base_url}/sparkling/{duration}", timeout=180)
            return res.status_code == 200
        except Exception as e:
            print(f"[Device] Sparkling Error: {e}")
            return False
    
    def stop_all_devices(self):
        """Emergency Stop for Devices"""
        try:
            requests.get(f"{self.base_url}/stopAll", timeout=10)
            print("[Device] Sent STOP ALL command")
        except:
            pass


# ---------------------------------------------------------
# Task & Planning
# ---------------------------------------------------------

class TaskStatus(Enum):
    PENDING = auto()
    RUNNING = auto()
    COMPLETED = auto()
    FAILED = auto()


class Task:
    def __init__(self, task_id: str, cmd_code: int, params: Dict[int, int] = None, 
                 dependencies: List[str] = None, order_uuid: str = None, skippable: bool = False):
        self.task_id = task_id
        self.cmd_code = cmd_code
        self.params = params or {}
        self.dependencies = dependencies or []
        self.status = TaskStatus.PENDING
        self.order_uuid = order_uuid
        self.skippable = skippable
        
        # Meta info for logging
        self.menu_name = ""
        self.order_no = 0
        
        # Atomic Sequence: Chained Task ID
        self.chained_next_task_id = None
        
        # Device Actions
        self.pre_device_action = None 
        self.post_device_action = None
        
        # Pickup Notify info
        self.notify_pickup = None
        
        # Pickup slot management
        self.assigned_slot = 0
        
        # 병렬 처리 관련
        self.parallel_check_point = False  # 커피 메뉴에서 병렬 처리 기회 확인 지점
        self.is_coffee_wait = False  # 커피 추출 대기 태스크


class TaskPlanner:
    """단일 로봇용 태스크 플래너"""
    def __init__(self):
        self.task_counter = 0
        self.recipes = {}
        self.load_recipes()

    def load_recipes(self):
        try:
            with open(RECIPE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    self.recipes = {item['menu_code']: item for item in data}
                else:
                    self.recipes = data
            print(f"[Planner] Loaded {len(self.recipes)} recipes.")
            
            # Simulation: Override Durations
            if SIMULATION_MODE:
                print("[Planner] Simulation Mode: Overriding recipe durations to 1.5s")
                for code, r in self.recipes.items():
                    r['water_ext_time'] = 1.5 if r.get('water_ext_time', 0) > 0 else 0
                    r['ice_ext_time'] = 1.5 if r.get('ice_ext_time', 0) > 0 else 0
                    r['hotwater_ext_time'] = 1.5 if r.get('hotwater_ext_time', 0) > 0 else 0
                    r['coffee_ext_time'] = 1.5 if r.get('coffee_ext_time', 0) > 0 else 0
                    r['icecream_ext_time'] = 1.5 if r.get('icecream_ext_time', 0) > 0 else 0
                    r['sparkling_ext_time'] = 1.5 if r.get('sparkling_ext_time', 0) > 0 else 0
                    if 'syrups' in r:
                        for s in r['syrups']:
                             if isinstance(s, dict) and 'time' in s:
                                 s['time'] = 1.5

        except Exception as e:
            print(f"[Planner] Failed to load recipes: {e}")

    def _new_id(self):
        self.task_counter += 1
        return f"T{self.task_counter}"

    def get_recipe(self, menu_code: int) -> Optional[Dict]:
        return self.recipes.get(menu_code)

    def is_coffee_menu(self, menu_code: int) -> bool:
        """커피가 포함된 메뉴인지 확인"""
        recipe = self.get_recipe(menu_code)
        if not recipe:
            return False
        return recipe.get('coffee_ext_time', 0) > 0

    def plan_order(self, order: Dict, order_uuid: str) -> List[Task]:
        """주문을 태스크 리스트로 변환 (단일 로봇용)"""
        logger.info(f"PLN|STR|{order['order_no']}|{order['menu_code']}|{order_uuid}")
        tasks = []
        menu_code = order.get('menu_code')
        order_type = order.get('order_type', 'DINEIN')
        recipe = self.get_recipe(menu_code)
        
        if not recipe:
            print(f"[Planner] Recipe not found for menu_code {menu_code}")
            return []
        
        # Validation
        cup_num = recipe.get('cup_num', 1)
        if cup_num <= 0 or cup_num > 2:
            print(f"[Planner] Error: Invalid Cup Number ({cup_num}). Must be 1(Hot) or 2(Ice)")
            return []
        
        # Manual Mode: No tasks
        if system_mode == MODE_MANUAL:
            print("[Planner] Manual Mode - Skipping planning")
            return []
            
        # 통합 플래닝 - recipe 값 기반으로 동적 생성
        tasks = self._plan_order_unified(recipe, order_type, order, order_uuid)

        # Inject Meta Info
        for t in tasks:
            t.menu_name = order.get('menu_name', '')
            t.order_no = order.get('order_no', 0)

        return tasks

    def _plan_serve_sequence(self, tasks, order_type, last_task_id, order, recipe, order_uuid):
        """서빙 시퀀스 (Move -> Place -> Notify -> Home) - 픽업대 1개"""
        # Move Task
        t_move = Task(self._new_id(), CMD_PICKUP_MOVE, {}, dependencies=[last_task_id], order_uuid=order_uuid)
        tasks.append(t_move)
        
        # Place Task
        t_serve = Task(self._new_id(), CMD_PICKUP_PLACE, {}, dependencies=[t_move.task_id], order_uuid=order_uuid)
        t_move.chained_next_task_id = t_serve.task_id
        t_serve.notify_pickup = {'zone': 1, 'order_no': order.get('order_no', 0), 'menu_code': recipe.get('menu_code', 0)}
        tasks.append(t_serve)
        
        # Home (Optional)
        t_home = Task(self._new_id(), CMD_HOME, {}, dependencies=[t_serve.task_id], order_uuid=order_uuid, skippable=True)
        tasks.append(t_home)
        
        return t_home.task_id

    def _plan_syrup_sequence(self, tasks, recipe, prev_task_id, order_uuid=None):
        """시럽 스테이션 시퀀스 (121→122)
        시럽이 여러 개인 경우 순차 처리
        """
        syrups = recipe.get('syrups', [])
        
        if not syrups:
            return prev_task_id
        
        last_task_id = prev_task_id
        
        # 시럽 종류별로 순차 처리
        for i, syrup in enumerate(syrups):
            syrup_id = syrup.get('id', 1)
            syrup_time = syrup.get('time', 3)
            
            # 시럽 스테이션 이동 (121)
            t_move = Task(self._new_id(), CMD_SYRUP_MOVE, {REG_SYRUP_IDX: syrup_id}, 
                         dependencies=[last_task_id], order_uuid=order_uuid)
            t_move.post_device_action = {'type': 'syrup', 'params': {'code': syrup_id, 'time': syrup_time}}
            tasks.append(t_move)
            
            # 시럽 완료 (122)
            t_done = Task(self._new_id(), CMD_SYRUP_DONE, {}, 
                         dependencies=[t_move.task_id], order_uuid=order_uuid)
            t_move.chained_next_task_id = t_done.task_id
            tasks.append(t_done)
            last_task_id = t_done.task_id
        
        return last_task_id

    # =========================================================
    # 통합 플래닝 (recipe 값 기반 동적 생성)
    # =========================================================
    def _plan_order_unified(self, recipe, order_type, order, order_uuid):
        """통합 플래닝 - recipe 값에 따라 필요한 단계만 동적으로 생성
        
        순서: 컵 → 제빙기(필요시) → 온수기(필요시) → 커피머신(필요시, 병렬체크) → 시럽(필요시) → 서빙
        """
        tasks = []
        
        # ===== 1. 컵 배출 (무조건) =====
        t_cup = Task(self._new_id(), CMD_CUP_MOVE, {REG_CUP_IDX: recipe['cup_num']}, order_uuid=order_uuid)
        tasks.append(t_cup)
        last_task_id = t_cup.task_id
        
        # ===== 2. 제빙기 (얼음/물/탄산수 필요시) =====
        ice_time = recipe.get('ice_ext_time', 0)
        water_time = recipe.get('water_ext_time', 0)
        sparkling_time = recipe.get('sparkling_ext_time', 0)
        
        if ice_time > 0 or water_time > 0 or sparkling_time > 0:
            t_wi_move = Task(self._new_id(), CMD_WI_MOVE, {}, dependencies=[last_task_id], order_uuid=order_uuid)
            t_wi_move.post_device_action = {
                'type': 'ice_water_sparkling',
                'params': {'ice': ice_time, 'water': water_time, 'sparkling': sparkling_time}
            }
            tasks.append(t_wi_move)
            
            # 얼음/물/탄산수 중 가장 큰 시간만큼 대기
            max_wait_time = max(ice_time, water_time, sparkling_time)
            
            t_wi_done = Task(self._new_id(), CMD_WI_DONE, {}, dependencies=[t_wi_move.task_id], order_uuid=order_uuid)
            t_wi_done.pre_device_action = {'type': 'sleep', 'params': {'time': max_wait_time}}  # CMD 112 전에 대기
            t_wi_move.chained_next_task_id = t_wi_done.task_id
            tasks.append(t_wi_done)
            last_task_id = t_wi_done.task_id
        
        # ===== 3. 온수기 (온수 필요시) =====
        hot_time = recipe.get('hotwater_ext_time', 0)
        
        if hot_time > 0:
            t_hot_move = Task(self._new_id(), CMD_HOT_MOVE, {}, dependencies=[last_task_id], order_uuid=order_uuid)
            t_hot_move.post_device_action = {'type': 'hot_water', 'params': {'time': hot_time}}  # 버튼 누름
            tasks.append(t_hot_move)
            
            t_hot_done = Task(self._new_id(), CMD_HOT_DONE, {}, dependencies=[t_hot_move.task_id], order_uuid=order_uuid)
            t_hot_done.pre_device_action = {'type': 'sleep', 'params': {'time': hot_time}}  # CMD 118 전에 대기
            t_hot_move.chained_next_task_id = t_hot_done.task_id
            tasks.append(t_hot_done)
            last_task_id = t_hot_done.task_id
        
        # ===== 4. 커피머신 (커피 필요시) - 병렬 처리 체크 포인트 =====
        coffee_time = recipe.get('coffee_ext_time', 0)
        
        if coffee_time > 0:
            coffee_product_id = recipe.get('coffee_product_id', 1)
            
            t_coffee_move = Task(self._new_id(), CMD_COFFEE_MOVE, {}, dependencies=[last_task_id], order_uuid=order_uuid)
            t_coffee_move.parallel_check_point = True
            
            # product_id == 1 (블랙 커피): 미리 그라인딩 시작 (pre_device_action)
            # product_id != 1 (우유 메뉴): 도착 후 추출 (post_device_action)
            coffee_action = {'type': 'coffee', 'params': {'id': coffee_product_id, 'time': 0.5}}
            if coffee_product_id == 1:
                t_coffee_move.pre_device_action = coffee_action
            else:
                t_coffee_move.post_device_action = coffee_action
            
            tasks.append(t_coffee_move)
            
            t_coffee_done = Task(self._new_id(), CMD_COFFEE_DONE, {}, dependencies=[t_coffee_move.task_id], order_uuid=order_uuid)
            t_coffee_done.is_coffee_wait = True
            t_coffee_done.pre_device_action = {'type': 'sleep', 'params': {'time': coffee_time}}  # CMD 114 보내기 전에 대기
            t_coffee_done.post_device_action = {'type': 'rinse'}  # CMD 114 완료 후 린스 강제 실행
            t_coffee_move.chained_next_task_id = t_coffee_done.task_id
            tasks.append(t_coffee_done)
            last_task_id = t_coffee_done.task_id
        
        # ===== 5. 시럽 스테이션 (시럽 필요시) - 서빙 직전 마지막 =====
        last_task_id = self._plan_syrup_sequence(tasks, recipe, last_task_id, order_uuid)
        
        # ===== 6. 서빙 (무조건) =====
        self._plan_serve_sequence(tasks, order_type, last_task_id, order, recipe, order_uuid)
        
        return tasks
    
# ---------------------------------------------------------
# Scheduler (단일 로봇 + 병렬 처리)
# ---------------------------------------------------------

class TaskScheduler:
    def __init__(self):
        self.tasks: List[Task] = []
        self.robot = RobotInterface('robot_1')  # 단일 로봇
        self.devices = DeviceInterface()
        
        # 픽업 슬롯 순환 (1→2→3→4→1...)
        self.next_pickup_slot = 1
        self.max_pickup_slots = 4
        
        self.running = False
        self.thread = None
        
        self.robot_busy = False
        self.robot_chained_task = None  # Atomic Sequence
        
        # 병렬 처리 상태
        self.parallel_mode = False
        self.parallel_completed = False
        self.paused_coffee_task = None     # 일시 중지된 커피 태스크
        self.paused_coffee_uuid = None     # 일시 중지된 커피 주문 UUID
        self.paused_coffee_order = None    # 일시 중지된 커피 주문 정보 (사본)
        self.coffee_start_time = 0         # 커피 추출 시작 시간
        self.coffee_duration = 0           # 커피 추출 시간
        
        # Coffee Rinse Logic
        self.coffeemachine_used = False
        self.last_coffee_time = time.time()  # 서버 시작 시간 기준으로 보상 로직 적용
        
        self.fail_safe_callback = None
        self.skip_callback = None
        self.status_callback = None
        self.order_manager = None
        self.planner = None
        self.session = requests.Session()

    def set_fail_safe_callback(self, callback):
        self.fail_safe_callback = callback

    def set_skip_condition_callback(self, callback):
        self.skip_callback = callback

    def set_status_callback(self, callback):
        self.status_callback = callback

    def set_order_manager(self, order_manager):
        self.order_manager = order_manager

    def set_planner(self, planner):
        self.planner = planner

    def get_pickup_slot(self) -> int:
        """픽업 슬롯 반환 (모드에 따라 동작)
        - rotate: 1→2→3→4→1... 순환 (센서 무시)
        - sensor: 센서 기반 빈 슬롯 탐색
        """
        if PICKUP_MODE == "sensor":
            return self._get_empty_pickup_slot_by_sensor()
        else:
            return self._get_next_pickup_slot_rotate()
    
    def _get_next_pickup_slot_rotate(self) -> int:
        """순환 방식으로 다음 픽업 슬롯 반환 (1→2→3→4→1...)"""
        slot = self.next_pickup_slot
        self.next_pickup_slot = (self.next_pickup_slot % self.max_pickup_slots) + 1
        print(f"[Scheduler] Pickup slot (rotate): {slot} (next: {self.next_pickup_slot})")
        return slot
    
    def _get_empty_pickup_slot_by_sensor(self) -> int:
        """센서 기반으로 빈 슬롯 탐색"""
        try:
            url = f"{PICKUP_SERVICE_URL}/getPickupStatus/1"
            res = self.session.get(url, timeout=10.0)
            if res.status_code == 200:
                data = res.json()
                status_arr = data.get('status', [])
                for i, occupied in enumerate(status_arr):
                    if occupied == 0:
                        print(f"[Scheduler] Pickup slot (sensor): {i + 1}")
                        return i + 1
        except Exception as e:
            print(f"[Scheduler] Failed to check pickup status: {e}")
        # 센서 실패 시 순환 모드로 폴백-> 센서 실패 시 상위루프에서 대기.
        print("[Scheduler] Pickup is FULL (No empty slot found)")
        return 0
    
    def reset_pickup_slot(self):
        """픽업 슬롯 순환 초기화"""
        self.next_pickup_slot = 1
        print("[Scheduler] Pickup slot counter reset to 1")

    def cancel_tasks(self, order_uuid: str):
        """Cancel PENDING tasks for a specific order UUID"""
        original_count = len(self.tasks)
        self.tasks = [t for t in self.tasks if t.order_uuid != order_uuid]
        removed_count = original_count - len(self.tasks)
        print(f"[Scheduler] Removed {removed_count} tasks for Order {order_uuid}")

    def start(self):
        self.running = True
        self.thread = threading.Thread(target=self._loop, daemon=True)
        self.thread.start()
        print("[Scheduler] Started")

    def stop_all(self):
        """Emergency Stop"""
        print("\[Scheduler] EMERGENCY STOP TRIGGERED")

        # Stop Robot Program using the correct API call
        try:
            print("\[Scheduler] Stopping robot program via API...")
            # 'set_system_mode'에서 사용하는 것과 동일한 API를 호출합니다.
            requests.get(f"{ROBOT_SERVICE_URL}/command/robot_1/1", timeout=5) # cmd 1 = stop_program
            logger.info("SYS|ROBOT_PROGRAM_STOP")
        except Exception as e:
            logger.error(f"SYS|ROBOT_PROGRAM_STOP_FAIL|{e}")
            print(f"\[Scheduler] Failed to stop robot program via API: {e}")

        # Stop Robot Motion as a secondary safety measure
        try:
            print("\[Scheduler] Stopping robot motion via register...")
            self.robot.write_register(REG_CMD, 6)  # Stop Motion
        except Exception as e:
            print(f"\[Scheduler] Failed to stop robot motion: {e}")
            
        # Clear Tasks
        self.tasks.clear()

        # Stop Devices
        self.devices.stop_all_devices()

        # Reset Flags
        self.robot_busy = False
        self.robot_chained_task = None
        self.parallel_mode = False
        self.parallel_completed = False
        self.paused_coffee_task = None
        self.paused_coffee_uuid = None
        self.paused_coffee_order = None

    def add_tasks(self, new_tasks: List[Task]):
        self.tasks.extend(new_tasks)
        print(f"[Scheduler] Added {len(new_tasks)} tasks. Total: {len(self.tasks)}")

    def _check_parallel_opportunity(self, current_order_uuid) -> Optional[str]:
        if not self.order_manager or not self.planner:
            return None
        
        waiting_orders = [
            o for o in self.order_manager.active_orders.values()
            if o['status'] == ORDER_WAITING and o['uuid'] != current_order_uuid
        ]
        
        # 주문 시간 순으로 정렬 (선입선출)
        waiting_orders.sort(key=lambda x: x.get('created_at', 0))
        
        for order in waiting_orders:
            # parallel_skip 플래그 체크 (이번 세션에서 실패한 주문 스킵)
            if order.get('parallel_skip'):
                continue
                
            menu_code = order.get('menu_code')
            if not self.planner.is_coffee_menu(menu_code):
                # 병렬 처리 대상 선택 → 상태 변경 및 기존 태스크 취소
                order['status'] = ORDER_PROCESSING
                self.cancel_tasks(order['uuid'])  # OrderManager가 생성한 태스크 취소
                return order['uuid']
        return None

    def _loop(self):
        while self.running:
            # 자동 린스 로직 제거 - CMD_COFFEE_DONE(114), CMD_COFFEE_PICK(116) 완료 후에만 린스 실행
            
            pending_tasks = [t for t in self.tasks if t.status == TaskStatus.PENDING]
            
            for task in pending_tasks:
                if self.robot_chained_task and task.task_id != self.robot_chained_task:
                    continue
                
                if self.robot_busy:
                    continue
                if not self._check_dependencies(task):
                        continue 
                
                threading.Thread(target=self._execute_task_wrapper, args=(task,)).start()
                break 

            time.sleep(0.1)

    def _check_dependencies(self, task: Task) -> bool:
        for dep_id in task.dependencies:
            dep_task = next((t for t in self.tasks if t.task_id == dep_id), None)
            if not dep_task or dep_task.status != TaskStatus.COMPLETED:
                return False
        return True

    def _execute_task_wrapper(self, task: Task):
        self.robot_busy = True
        notify_clients('robot_updated')
        
        try:
            task.status = TaskStatus.RUNNING
            
            if task.order_uuid and self.status_callback:
                self.status_callback(task.order_uuid, ORDER_PROCESSING)

            should_skip = False
            if task.skippable:
                if self.skip_callback and self.skip_callback():
                    should_skip = True
                else:
                    pending_count = len([t for t in self.tasks if t.status == TaskStatus.PENDING])
                    if pending_count > 0:
                         should_skip = True

            if should_skip:
                print(f"[Scheduler] Skipping task {task.task_id}")
                self.robot_chained_task = None
            else:
                self._execute_task(task)
                
                if self.parallel_completed:
                    self.robot_chained_task = None
                    self.parallel_completed = False
                    print(f"[Scheduler] Parallel mode completed. Chain reset.")
                elif task.chained_next_task_id:
                    self.robot_chained_task = task.chained_next_task_id
                else:
                    self.robot_chained_task = None
            
            task.status = TaskStatus.COMPLETED
            
            if task.order_uuid and self.status_callback:
                remaining = [t for t in self.tasks if t.order_uuid == task.order_uuid and t.status != TaskStatus.COMPLETED]
                if not remaining:
                    self.status_callback(task.order_uuid, ORDER_COMPLETED)
            
        except Exception as e:
            msg = f"[Scheduler][Error] Task {task.task_id} failed: {e}"
            print(msg)
            logger.error(msg)
            task.status = TaskStatus.FAILED
            self.robot_chained_task = None
            
            if "Cup Dispense Failed" in str(e) or "Timeout" in str(e):
                print(f"[Scheduler] Critical Error. Initiating Fail Safe...")
                if self.fail_safe_callback:
                    self.fail_safe_callback()
            
        finally:
            self.robot_busy = False
            notify_clients('robot_updated')

    def _execute_task(self, task: Task):
        robot = self.robot
        
        # ═══════════════════════════════════════════════════════════════
        # 실행할 명령 코드 결정 (113 vs 115)
        # ═══════════════════════════════════════════════════════════════
        actual_cmd = task.cmd_code  # 기본값: 태스크에 지정된 명령
        parallel_uuid = None
        
        if task.parallel_check_point and self.order_manager and self.planner:
            parallel_uuid = self._check_parallel_opportunity(task.order_uuid)
            
            if parallel_uuid:
                # 비커피 주문 발견 → 115(Place)로 변경
                actual_cmd = CMD_COFFEE_PLACE
                print(f"[Parallel] Non-coffee order found: {parallel_uuid}")
                print(f"[Parallel] Command changed: {task.cmd_code}(MOVE) → {actual_cmd}(PLACE)")
                
                self.paused_coffee_task = task
                self.paused_coffee_uuid = task.order_uuid
                coffee_order = self.order_manager.active_orders.get(task.order_uuid)
                if coffee_order:
                    self.paused_coffee_order = coffee_order.copy()  # 사본 저장
                print(f"[Parallel] Saved paused coffee order: {self.paused_coffee_uuid}")
            else:
                print(f"[Scheduler] No parallel opportunity. Using normal mode: {task.cmd_code}")
        
        # ═══════════════════════════════════════════════════════════════
        # Pre Action (커피 추출 시작 등)
        # ═══════════════════════════════════════════════════════════════
        if task.pre_device_action:
            action = task.pre_device_action
            
            # 써모플랜 보일러 온도 보상 로직
            # 커피 대기 태스크이고, 5분 이상 커피머신 미사용 시 추가 대기 시간 적용
            if task.is_coffee_wait and COFFEE_BRAND == "thermoplan":
                time_since_last = time.time() - self.last_coffee_time
                if self.last_coffee_time > 0 and time_since_last > IDLE_TIME_THRESHOLD_SECONDS:
                    original_time = action.get('params', {}).get('time', 0)
                    extra_time = EXTRA_DURATION_SECONDS
                    new_time = original_time + extra_time
                    action = {
                        'type': action.get('type'),
                        'params': {'time': new_time}
                    }
                    print(f"[Thermoplan] Idle time ({time_since_last:.0f}s) exceeded threshold. Adding {extra_time}s to duration. ({original_time}s -> {new_time}s)")
            
            self._execute_device_action(action)

        # ═══════════════════════════════════════════════════════════════
        # Pickup Slot Assignment (모드에 따라: sensor 또는 rotate)
        # ═══════════════════════════════════════════════════════════════
        if task.cmd_code == CMD_PICKUP_PLACE:
            if PICKUP_MODE == "sensor":
                # 센서 모드: 빈 슬롯 나올 때까지 대기
                while True:
                    slot = self.get_pickup_slot()
                    if slot > 0:
                        break
                    print("[Scheduler] Pickup is FULL. Waiting...")
                    time.sleep(2.0)
                    if not self.running:
                        return
            else:
                # 순환 모드: 대기 없이 바로 할당
                slot = self.get_pickup_slot()
            
            task.assigned_slot = slot
            task.params[REG_F_PICK_IDX] = slot
            print(f"[Scheduler] Assigned Pickup Slot: {slot} (mode: {PICKUP_MODE})")

        # ═══════════════════════════════════════════════════════════════
        # 로봇 명령 실행
        # ═══════════════════════════════════════════════════════════════
        cmd_name = CMD_DESC.get(actual_cmd, "UNK")
        logger.info(f"TSK|STR|{task.task_id}|{actual_cmd}|{cmd_name}|{task.order_no}|{task.menu_name}")
        
        current_init = robot.read_register(REG_INIT)
        if current_init != 0:
            robot.write_register(REG_INIT, 0)
            time.sleep(0.5)

        for addr, val in task.params.items():
            robot.write_register(addr, val)
            time.sleep(0.05)
            
        # Send Command
        robot.send_command(actual_cmd)
        time.sleep(0.5)
        
        expected_init = actual_cmd + 500
        
        # CMD_CUP_MOVE(110)는 컵 추출 과정 후에 wait_init
        if task.cmd_code == CMD_CUP_MOVE:
            # ─────────────────────────────────────────────────────────────
            # 컵 배출 프로세스
            # ─────────────────────────────────────────────────────────────
            # 1. CUP_ON(106) 대기 - 로봇이 컵 디스펜서에 도착하면 1
            print("[Cup] Waiting for CUP_ON (106)...")
            cup_on_timeout = 60.0
            t0 = time.time()
            while time.time() - t0 < cup_on_timeout:
                cup_on = robot.read_register(REG_CUP_ON)
                print('cup on : ', cup_on)
                if cup_on == 1:
                    break
                time.sleep(0.1)
            else:
                raise Exception("Cup Dispense Timeout: CUP_ON not received")
            
            # 2. CUP_ON(106) 초기화
            robot.write_register(REG_CUP_ON, 0)
            print("[Cup] CUP_ON received, reset to 0")
            
            # 3. 컵 추출 신호 보냄 (DO 5번 카드, HOT=3번→3203 / ICE=4번→3204)
            cup_idx = task.params.get(REG_CUP_IDX, 1)  # 1: HOT, 2: ICE
            coil_addr = 3202 if cup_idx == 1 else 3203  # HOT=5/3, ICE=5/4
            try:
                requests.get(f"{IO_SERVICE_URL}/coil/write/5/{coil_addr}/1", timeout=5)
                print(f"[Cup] Dispense signal sent (Unit=5, Addr={coil_addr}, Value=1)")
                time.sleep(1.0)  # 신호 유지
                requests.get(f"{IO_SERVICE_URL}/coil/write/5/{coil_addr}/0", timeout=5)
                print(f"[Cup] Dispense signal off (Unit=5, Addr={coil_addr}, Value=0)")
            except Exception as e:
                print(f"[Cup] IO Error: {e}")
            
            cup_index_idx = 3 if cup_idx == 1 else 4
            robot.write_register(REG_CUP_IDX, cup_index_idx)
            
            # REG_CUP_SENSOR / REG_CUP_MOVE
            # 로봇 -> main 센서 위치 도착 (104 - 1 성공)
            # 105번에 센서인식o:1, 센서인식x:2
            # 4. CUP_RES(102) 대기 - 로봇이 센서 체크 후 결과 기록
            print("[CUP] waiting for CUP_MOVE (104")
            cup_move_timeout = 60.0
            t1 = time.time()
            while time.time() - t1 < cup_move_timeout:
                res = robot.read_register(REG_CUP_MOVE)
                if int(res) == 1:
                    break
                time.sleep(0.1)
            else:
                raise Exception("Cup Dispense Timeout : CUP_MOVE")
            
            # REG_CUP_MOVE 초기화
            robot.write_register(REG_CUP_MOVE, 0)
            print("[Cup] CUP_MOVE received, reset to 0")
            
            # di 수정 필요
            # 센서 인식 (DI 3번 카드, 3206번 주소)
            cup_sensor_ok = False
            try:
                res = requests.get(f"{IO_SERVICE_URL}/coils/read/3/6/1", timeout=5)
                sensor_data = res.json()
                print(f"[Cup] Sensor read result: {sensor_data}")
                
                # 센서 값 확인 (1: 감지됨, 0: 감지 안됨)
                if sensor_data and len(sensor_data) > 0 and sensor_data[0] == 1:
                    robot.write_register(REG_CUP_SENSOR, 1)  # 센싱 O
                    print("[Cup] Sensor detected -> REG_CUP_SENSOR = 1")
                    cup_sensor_ok = True
                else:
                    robot.write_register(REG_CUP_SENSOR, 2)  # 센싱 X
                    print("[Cup] Sensor NOT detected -> REG_CUP_SENSOR = 2")
            except Exception as e:
                print(f"[Cup] Sensor read error: {e}")
                robot.write_register(REG_CUP_SENSOR, 2)  # 에러시 센싱 X로 처리
            
            # 컵 센서 실패 시: 로봇 홈 대기 → 주문 종료 → 수동 모드 전환
            if not cup_sensor_ok:
                print("\[Cup] Cup sensor failed! Waiting for robot to go home...")

                # 1. 로봇 홈 대기 (610 기다림) - 수동 모드 전환 전에!
                if not robot.wait_init(expected_init, timeout=600.0):
                    print("\[Cup] Robot Init Timeout after Cup Fail")
                robot.write_register(REG_INIT, 0)

                # 2. 주문 완료 처리
                if self.status_callback and task.order_uuid:
                    self.status_callback(task.order_uuid, ORDER_COMPLETED)

                # 3. Fail Safe 핸들러를 호출하여 시스템을 안전하게 정지
                if self.fail_safe_callback:
                    print("\[Cup] Cup sensor failed. Initiating Fail Safe...")
                    self.fail_safe_callback()
                    

                # 4. 현재 진행중인 태스크를 중단하기 위해 예외 발생
                raise Exception("Cup Sensor Failed: Switched to Manual Mode")

            #cup_sensor_ok = True  # 임시: 항상 성공으로 처리
            #robot.write_register(REG_CUP_SENSOR, 1)  # 임시: 항상 1 고정
            #print("[Cup] Sensor check SKIPPED (forced to 1 for testing)")
            
            # 컵 추출 완료 후 wait_init(610)
            if not robot.wait_init(expected_init, timeout=600.0):
                raise Exception("Robot Init Timeout after Cup Dispense")
            robot.write_register(REG_INIT, 0)
            print(f"[Cup] Robot completed (Init={expected_init})")
        else:
            # CMD_CUP_MOVE가 아닌 경우 기존 wait_init 로직
            if not robot.wait_init(expected_init, timeout=600.0):
                raise Exception("Robot Init Timeout")
            robot.write_register(REG_INIT, 0)

        # ═══════════════════════════════════════════════════════════════
        # 병렬 처리 모드 (615 확인 후)
        # ═══════════════════════════════════════════════════════════════
        if expected_init == 615:
            # wait_init(615) 완료 후 바로 병렬 처리 시작
            robot.write_register(REG_INIT, 0)  # 700번 리셋
            print(f"[Parallel] Init 615 received and reset, starting parallel processing")
            
            self._process_parallel_order(task, parallel_uuid)
            return 
            
        # ═══════════════════════════════════════════════════════════════
        # Post Action (일반 모드)
        # ═══════════════════════════════════════════════════════════════
        if task.post_device_action:
            self._execute_device_action(task.post_device_action)
            
        # ═══════════════════════════════════════════════════════════════
        # Notify Pickup
        # ═══════════════════════════════════════════════════════════════
        if task.notify_pickup:
            self._notify_pickup_service(task.notify_pickup, task.assigned_slot)
            self._auto_clear_pickup_sim(task.notify_pickup['zone'], task.assigned_slot)
            
        logger.info(f"TSK|END|{task.task_id}")

    def _process_parallel_order(self, coffee_task: Task, parallel_uuid: str):
        """
        병렬 처리: 비커피 음료 제조 + 커피 Pick
        
        1. 커피 추출 명령 전송 (post_device_action이 있는 경우, product_id != 1)
        2. 커피 추출 시간 기록
        3. 비커피 음료 제조 및 서빙 (반복)
           - 남은 시간 ≥ 20초: 추가 비커피 확인 후 처리
           - 남은 시간 < 20초: 대기
        4. 커피 추출 완료 대기
        5. 116(Pick) 실행
        6. CMD_COFFEE_DONE 태스크 건너뛰기 처리
        7. 병렬 모드 종료 및 상태 초기화
        
        기존 커피 주문 정보:
        - self.paused_coffee_task: 일시 중지된 커피 태스크
        - self.paused_coffee_uuid: 일시 중지된 커피 주문 UUID
        - self.paused_coffee_order: 일시 중지된 커피 주문 정보 (사본)
        """
        robot = self.robot
        
        self.parallel_mode = True
        
        PARALLEL_THRESHOLD = 20
        
        print(f"[Parallel] === Starting Parallel Processing ===")
        print(f"[Parallel] Paused Coffee UUID: {self.paused_coffee_uuid}")
        print(f"[Parallel] Paused Coffee Order: {self.paused_coffee_order.get('menu_name') if self.paused_coffee_order else 'N/A'}")
        
        # ─────────────────────────────────────────────────────────────
        # 1. 커피 추출 명령 전송 (post_device_action이 있는 경우)
        #    - product_id == 1: pre_device_action에서 이미 실행됨
        #    - product_id != 1: post_device_action이 있으므로 여기서 실행
        # ─────────────────────────────────────────────────────────────
        if coffee_task.post_device_action:
            action = coffee_task.post_device_action
            if action.get('type') == 'coffee':
                print(f"[Parallel] Executing post_device_action (coffee): {action}")
                self._execute_device_action(action)
        
        # ─────────────────────────────────────────────────────────────
        # 2. 커피 추출 시간 기록 (저장된 주문 정보 사용) + 보일러 보상
        # ─────────────────────────────────────────────────────────────
        recipe = None
        if self.paused_coffee_order and self.planner:
            recipe = self.planner.get_recipe(self.paused_coffee_order.get('menu_code'))
        
        coffee_duration = recipe.get('coffee_ext_time', 30) if recipe else 30
        
        # 써모플랜 보일러 온도 보상 로직 (병렬 처리에서도 적용)
        if COFFEE_BRAND == "thermoplan" and self.last_coffee_time > 0:
            time_since_last = time.time() - self.last_coffee_time
            if time_since_last > IDLE_TIME_THRESHOLD_SECONDS:
                original_duration = coffee_duration
                coffee_duration += EXTRA_DURATION_SECONDS
                print(f"[Parallel][Thermoplan] Idle time ({time_since_last:.0f}s) exceeded threshold. Adding {EXTRA_DURATION_SECONDS}s to duration. ({original_duration}s -> {coffee_duration}s)")
        
        coffee_start_time = time.time()
        
        print(f"[Parallel] Coffee extraction started. Duration: {coffee_duration}s")
        
        # ─────────────────────────────────────────────────────────────
        # 3. 비커피 음료 제조 루프 (여러 개 연속 처리 가능)
        # ─────────────────────────────────────────────────────────────
        current_parallel_uuid = parallel_uuid
        parallel_count = 0
        
        while current_parallel_uuid and self.running:
            parallel_count += 1
            parallel_order = self.order_manager.active_orders.get(current_parallel_uuid)
            
            if not parallel_order:
                break
            
            # 병렬 주문 상태를 즉시 PROCESSING으로 변경 (중복 처리 방지)
            parallel_order['status'] = ORDER_PROCESSING
                
            print(f"[Parallel] Processing #{parallel_count}: {parallel_order.get('menu_name')} (UUID: {current_parallel_uuid})")
        
            parallel_tasks = self.planner.plan_order(parallel_order, current_parallel_uuid)
        
            for t in parallel_tasks:
                t.menu_name = parallel_order.get('menu_name', '')
                t.order_no = parallel_order.get('order_no', 0)
            
            for pt in parallel_tasks:
                if not self.running:
                    break
                
                # 병렬 처리 중에는 skippable 태스크(HOME) 스킵 - 커피 픽업하러 가야 함
                if pt.skippable:
                    print(f"[Parallel] Skipping HOME task: {pt.task_id} (CMD {pt.cmd_code})")
                    pt.status = TaskStatus.COMPLETED
                    continue
                    
                pt.status = TaskStatus.RUNNING
                
                if self.status_callback:
                    self.status_callback(current_parallel_uuid, ORDER_PROCESSING)
                
                try:
                    self._execute_task(pt)
                    pt.status = TaskStatus.COMPLETED
                except Exception as e:
                    print(f"[Parallel] Task {pt.task_id} failed: {e}")
                    pt.status = TaskStatus.FAILED
                    
                    # 실패한 병렬 주문 복구 및 스킵 플래그 설정
                    parallel_order['status'] = ORDER_WAITING  # 상태 복구
                    parallel_order['parallel_skip'] = True    # 이번 세션에서 스킵
                    print(f"[Parallel] Order restored to WAITING with parallel_skip=True")
                    break
            
            # 실패하지 않은 경우에만 COMPLETED 설정
            if parallel_tasks and pt.status != TaskStatus.FAILED:
                if self.status_callback:
                    self.status_callback(current_parallel_uuid, ORDER_COMPLETED)
            
            print(f"[Parallel] Parallel order #{parallel_count} completed")
            
            # ─────────────────────────────────────────────────────────
            # 남은 시간 체크 → 추가 비커피 처리 여부 결정
            # ─────────────────────────────────────────────────────────
            elapsed = time.time() - coffee_start_time
            remaining = coffee_duration - elapsed
            
            print(f"[Parallel] Coffee remaining time: {remaining:.1f}s")
            
            if remaining >= PARALLEL_THRESHOLD:
                next_parallel = self._check_parallel_opportunity(self.paused_coffee_uuid)
                if next_parallel:
                    print(f"[Parallel] {remaining:.1f}s remaining >= {PARALLEL_THRESHOLD}s. Found another non-coffee order.")
                    current_parallel_uuid = next_parallel
                    continue
                else:
                    print(f"[Parallel] {remaining:.1f}s remaining >= {PARALLEL_THRESHOLD}s. No more non-coffee orders.")
                    current_parallel_uuid = None
            else:
                print(f"[Parallel] {remaining:.1f}s remaining < {PARALLEL_THRESHOLD}s. Stopping parallel processing.")
                current_parallel_uuid = None
        
        print(f"[Parallel] Total parallel orders processed: {parallel_count}")
        
        # ─────────────────────────────────────────────────────────────
        # 4. 커피 추출 완료 대기
        # ─────────────────────────────────────────────────────────────
        elapsed = time.time() - coffee_start_time
        remaining = coffee_duration - elapsed
        
        if remaining > 0:
            print(f"[Parallel] Waiting for coffee: {remaining:.1f}s remaining")
            time.sleep(remaining)
        
        print(f"[Parallel] Coffee ready. Picking up...")
        
        # ─────────────────────────────────────────────────────────────
        # 5. 커피머신에서 Pick (116)
        # ─────────────────────────────────────────────────────────────
        robot.write_register(REG_INIT, 0)
        time.sleep(0.3)
        
        robot.send_command(CMD_COFFEE_PICK)  # 116
        time.sleep(0.5)
        
        expected_init = CMD_COFFEE_PICK + 500  # 616
        if not robot.wait_init(expected_init, timeout=600.0):
            raise Exception("Robot Init Timeout (Parallel Coffee Pick)")
        robot.write_register(REG_INIT, 0)
        
        logger.info(f"TSK|PARALLEL_PICK|{coffee_task.task_id}|116")
            
        self.coffeemachine_used = True
        self.last_coffee_time = time.time()
        
        # CMD 116 완료 후 린스 강제 실행
        print("[Parallel] Executing rinse after coffee pick...")
        threading.Thread(target=self.devices.execute_rinse, daemon=True).start()
        
        # ─────────────────────────────────────────────────────────────
        # 6. CMD_COFFEE_DONE(114) 태스크 건너뛰기 처리
        # ─────────────────────────────────────────────────────────────
        if coffee_task.chained_next_task_id:
            coffee_done_task = next((t for t in self.tasks if t.task_id == coffee_task.chained_next_task_id), None)
            if coffee_done_task:
                print(f"[Parallel] Skipping CMD_COFFEE_DONE: {coffee_done_task.task_id}")
                coffee_done_task.status = TaskStatus.COMPLETED
        
        # ─────────────────────────────────────────────────────────────
        # 7. 병렬 모드 종료 및 상태 초기화
        # ─────────────────────────────────────────────────────────────
        print(f"[Parallel] Resuming coffee order: {self.paused_coffee_uuid}")
        print(f"[Parallel] Coffee menu: {self.paused_coffee_order.get('menu_name') if self.paused_coffee_order else 'N/A'}")
        
        self.parallel_completed = True
        self.parallel_mode = False
        
        self.paused_coffee_task = None
        self.paused_coffee_uuid = None
        self.paused_coffee_order = None
        
        # parallel_skip 플래그 초기화 (다음 세션에서 다시 병렬 처리 가능)
        for order in self.order_manager.active_orders.values():
            order.pop('parallel_skip', None)
        print(f"[Parallel] Cleared parallel_skip flags")
        
        print(f"[Parallel] === Parallel Processing Completed ===")

    def _notify_pickup_service(self, info, slot):
        zone = info['zone']
        order_no = info['order_no']
        menu_code = info['menu_code']
        
        if slot == 0:
            print("[Scheduler] Warning: No slot assigned for pickup notification")
            return
            
        try:
            url = f"{PICKUP_SERVICE_URL}/updateDID/{zone}/{slot}/{order_no}/{menu_code}"
            self.session.get(url, timeout=10.0)
            logger.debug(f"[Scheduler] Notified Pickup: Zone {zone} Slot {slot} Order {order_no}")
        except Exception as e:
            logger.error(f"[Scheduler] Failed to notify Pickup: {e}")

    def _auto_clear_pickup_sim(self, zone_id, slot):
        if not SIMULATION_MODE: return
        
        def _clear():
            time.sleep(2.0) 
            try:
                    requests.get(f"{IO_SERVICE_URL}/sim/setPickup/{zone_id}/{slot}/0", timeout=1)
            except:
                pass
        
        threading.Thread(target=_clear, daemon=True).start()

    def _execute_device_action(self, action):
        act_type = action.get('type')
        p = action.get('params', {})
        logger.info(f"DEV|{act_type}")
        
        success = False
        
        if act_type == 'coffee':
            success = self.devices.make_coffee(p.get('id'), p.get('time'))
            if success:
                self.coffeemachine_used = True
                # last_coffee_time은 커피 추출 완료(린스) 후에만 업데이트
                # 여기서 업데이트하면 보일러 보상 로직이 작동하지 않음     
        elif act_type == 'ice_water':
            success = self.devices.dispense_ice_water(p.get('ice'), p.get('water')) 
        elif act_type == 'ice_water_sparkling':
            success = self.devices.dispense_ice_water(p.get('ice'), p.get('water'))
            if success and p.get('sparkling', 0) > 0:
                success = self.devices.dispense_sparkling(p.get('sparkling'))  
        elif act_type == 'hot_water':
            success = self.devices.dispense_hot_water(p.get('time'))
        elif act_type == 'syrup':
            success = self.devices.dispense_syrup(p.get('code'), p.get('time'))
        elif act_type == 'sparkling':
            success = self.devices.dispense_sparkling(p.get('time'))
        elif act_type == 'sleep':
            try:
                duration = float(p.get('time', 0))
                if SIMULATION_MODE:
                    duration = 1.5
                print(f"[DeviceAction] Sleeping for {duration}s...")
                time.sleep(duration)
                success = True
            except Exception as e:
                print(f"[DeviceAction] Sleep Error: {e}")
                success = False
        elif act_type == 'rinse':
            # 커피 추출 완료 후 린스 강제 실행 (비동기)
            print("[DeviceAction] Executing rinse after coffee done...")
            threading.Thread(target=self.devices.execute_rinse, daemon=True).start()
            self.coffeemachine_used = False  # 린스 실행했으므로 플래그 리셋
            self.last_coffee_time = time.time()  # 린스도 커피머신 사용이므로 시간 기록
            success = True
            
        if not success:
            raise Exception(f"Device Action {act_type} Failed")


# ---------------------------------------------------------
# Order Management
# ---------------------------------------------------------

class OrderManager:
    def __init__(self, planner: TaskPlanner, scheduler: TaskScheduler):
        self.order_queue = Queue()
        self.active_orders = {} 
        self.planner = planner
        self.scheduler = scheduler
        self.scheduler.set_status_callback(self.update_order_status)
        self.scheduler.set_skip_condition_callback(lambda: not self.order_queue.empty())
        self.scheduler.set_order_manager(self)
        self.scheduler.set_planner(planner)
        self.running = True
        self.thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self.thread.start()
        print("[OrderManager] Started")

    def add_order(self, order):
        order_uuid = f"{int(time.time() * 1000)}"
        order['uuid'] = order_uuid
        order['status'] = ORDER_WAITING
        order['created_at'] = time.time()
        
        self.active_orders[order_uuid] = order
        self.order_queue.put(order_uuid)
        
        print(f"[OrderManager] Order Added: {order_uuid} ({order.get('menu_name', '')})")
        logger.info(f"ORD|ADD|{order['order_no']}|{order['menu_code']}|{order.get('menu_name','')}|{order_uuid}")
        
        notify_clients('order_updated')
        return order_uuid

    def update_order_status(self, order_uuid, status):
        if order_uuid in self.active_orders:
            old_status = self.active_orders[order_uuid]['status']
            self.active_orders[order_uuid]['status'] = status
            
            if status == ORDER_COMPLETED:
                self.active_orders[order_uuid]['completed_at'] = time.time()
                log_performance_to_excel(self.active_orders[order_uuid])
                logger.info(f"ORD|CMP|{self.active_orders[order_uuid]['order_no']}|{order_uuid}")
                del self.active_orders[order_uuid]  # 완료 후 삭제
            
            elif status == ORDER_CANCELLED:
                del self.active_orders[order_uuid]  # 취소 후 삭제
            
            notify_clients('order_updated')

    def cancel_order(self, order_uuid):
        if order_uuid in self.active_orders:
            self.active_orders[order_uuid]['status'] = ORDER_CANCELLED
            self.scheduler.cancel_tasks(order_uuid)
            notify_clients('order_updated')
            return True
        return False

    def _monitor_loop(self):
        while self.running:
            if system_mode != MODE_AUTO:
                time.sleep(0.5)
                continue
                
            try:
                order_uuid = self.order_queue.get(timeout=1.0)
                if order_uuid not in self.active_orders:
                    continue
            
                order = self.active_orders[order_uuid]
                if order['status'] != ORDER_WAITING:
                    continue
            
                tasks = self.planner.plan_order(order, order_uuid)
                if tasks:
                    self.scheduler.add_tasks(tasks)
                else:
                    pass
            except Empty:
                pass
            except Exception as e:
                print(f"[OrderManager] Error: {e}")
                time.sleep(1.0)


# ---------------------------------------------------------
# Global State & API
# ---------------------------------------------------------
system_mode = MODE_MANUAL
planner = None
scheduler = None
order_manager = None

app = Flask(__name__)
CORS(app)

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'OK', 'mode': system_mode})

@app.route('/getSystemMode', methods=['GET'])
def get_system_mode():
    return jsonify({'mode': system_mode})

@app.route('/setSystemMode/<int:mode>', methods=['GET', 'POST'])
def set_system_mode(mode):
    global system_mode
    old_mode = system_mode
    system_mode = mode
    
    if mode == MODE_MANUAL and old_mode == MODE_AUTO:
        scheduler.stop_all()
        # 로봇 프로그램 정지
        try:
            requests.get(f"{ROBOT_SERVICE_URL}/command/robot_1/1", timeout=5)  # cmd 1 = stop_program
            logger.info("SYS|ROBOT_PROGRAM_STOP")
        except Exception as e:
            logger.error(f"SYS|ROBOT_PROGRAM_STOP_FAIL|{e}")
        try:
            requests.get(f"{PICKUP_SERVICE_URL}/resetAll", timeout=5)
        except:
            pass
    
    # 자동 모드 시작 시 로봇 프로그램 1번 실행 + 픽업 슬롯 초기화
    if mode == MODE_AUTO and old_mode == MODE_MANUAL:
        scheduler.reset_pickup_slot()
        try:
            requests.get(f"{ROBOT_SERVICE_URL}/runProgram/robot_1/1", timeout=5)
            logger.info("SYS|ROBOT_PROGRAM_START|1")
        except Exception as e:
            logger.error(f"SYS|ROBOT_PROGRAM_START_FAIL|{e}")
    
    logger.info(f"SYS|MODE|{old_mode}|->|{mode}")
    notify_clients('system_mode_changed', {'mode': mode})
    return jsonify({'mode': system_mode})

@app.route('/addOrder/<int:order_no>/<int:menu_code>', methods=['GET'])
def add_order_url(order_no, menu_code):
    recipe = planner.get_recipe(menu_code) if planner else None
    menu_name = recipe.get('menu_name', f'Menu {menu_code}') if recipe else f'Menu {menu_code}'

    order = {
        'order_no': order_no,
        'menu_code': menu_code,
        'menu_name': menu_name
    }
    
    order_uuid = order_manager.add_order(order)
    return jsonify({'uuid': order_uuid, 'status': ORDER_WAITING})

@app.route('/addOrder', methods=['POST'])
def add_order_json():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400
    
    if 'menu_name' not in data and 'menu_code' in data:
        recipe = planner.get_recipe(data['menu_code']) if planner else None
        data['menu_name'] = recipe.get('menu_name', f"Menu {data['menu_code']}") if recipe else f"Menu {data['menu_code']}"
    
    order_uuid = order_manager.add_order(data)
    return jsonify({'uuid': order_uuid, 'status': ORDER_WAITING})
@app.route('/getOrders', methods=['GET'])
def get_orders():
    return jsonify(list(order_manager.active_orders.values()))

@app.route('/getActiveOrders', methods=['GET'])
def get_active_orders():
    active = {o['uuid']: o for o in order_manager.active_orders.values() 
              if o['status'] in [ORDER_WAITING, ORDER_PROCESSING]}
    return jsonify({'orders': active})

@app.route('/cancelOrder/<string:order_uuid>', methods=['GET', 'POST'])
def cancel_order(order_uuid):
    success = order_manager.cancel_order(order_uuid)
    return jsonify({'success': success})

@app.route('/getSchedulerStatus', methods=['GET'])
def get_scheduler_status():
    pending = len([t for t in scheduler.tasks if t.status == TaskStatus.PENDING])
    running = len([t for t in scheduler.tasks if t.status == TaskStatus.RUNNING])
    return jsonify({
        'pending_tasks': pending,
        'running_tasks': running,
        'robot_busy': scheduler.robot_busy,
        'parallel_mode': scheduler.parallel_mode
    })

@app.route('/emergencyStop', methods=['GET', 'POST'])
def emergency_stop():
    global system_mode
    system_mode = MODE_MANUAL
    scheduler.stop_all()
    logger.info("SYS|EMERGENCY_STOP")
    notify_clients('system_mode_changed', {'mode': MODE_MANUAL})
    return jsonify({'status': 'stopped'})

@app.route('/getAllRecipes', methods=['GET'])
def get_all_recipes():
    return jsonify(list(planner.recipes.values()))

@app.route('/getRecipe/<int:menu_code>', methods=['GET'])
def get_recipe(menu_code):
    recipe = planner.get_recipe(menu_code)
    if recipe:
        return jsonify(recipe)
    return jsonify({'error': 'Not found'}), 404

def fail_safe_handler():
    """Critical Error -> Switch to Manual Mode"""
    global system_mode
    print("[FailSafe] Switching to Manual Mode")
    system_mode = MODE_MANUAL
    scheduler.stop_all()
    notify_clients('system_mode_changed', {'mode': MODE_MANUAL})

def initialize():
    global planner, scheduler, order_manager
    
    planner = TaskPlanner()
    scheduler = TaskScheduler()
    scheduler.set_fail_safe_callback(fail_safe_handler)
    scheduler.start()
    
    order_manager = OrderManager(planner, scheduler)
    
    print("[System] Order Service Initialized (Single Robot Mode)")


if __name__ == '__main__':
    initialize()
    app.run(host='0.0.0.0', port=8100, debug=False, threaded=True)
